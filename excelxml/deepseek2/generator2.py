import pandas as pd
import xml.etree.ElementTree as ET
from xml.dom import minidom
from pathlib import Path

from openpyxl import load_workbook

# --------------------------
# 读取 Excel 文件
# --------------------------
def read_excel(file_path):
    df = pd.read_excel(file_path, engine='openpyxl')
    return df.to_dict(orient='records')  # 转换为字典列表

# --------------------------
# 生成 XML 文件
# --------------------------
def generate_xml(data, output_path):
    # 创建根节点
    root = ET.Element("Data")
    
    # 添加每条记录
    for item in data:
        record = ET.SubElement(root, "Record")
        for key, value in item.items():
            field = ET.SubElement(record, key)
            field.text = str(value)
    
    # 美化 XML 格式
    xml_str = minidom.parseString(ET.tostring(root)).toprettyxml(indent="  ")
    
    # 写入文件
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(xml_str)

# --------------------------
# 写入新 Excel 文件（示例：添加处理后的数据）
# --------------------------
def write_excel(data, output_path):
    # 示例：在原数据基础上新增一列
    df = pd.DataFrame(data)
    df["Processed"] = "已处理"  # 添加新列
    
    # 写入 Excel
    df.to_excel(output_path, index=False, engine='openpyxl')

# --------------------------
# 追加写入 Excel 文件（示例：添加处理后的数据）
# --------------------------
def append_excel(data, output_path):
    # 加载工作簿并选择活动 Sheet
    wb = load_workbook(output_path)
    ws = wb.active

    # 找到最后一行的下一行
    next_row = ws.max_row + 1
    for item in data:
        ws.append(list(item.values()))

    wb.save(output_path)
    print("追加数据完成")


# --------------------------
# 主程序
# --------------------------
if __name__ == "__main__":
    # 文件路径配置
    input_excel = Path(__file__).parent/'input.xlsx';
    output_xml = "output.xml"
    output_excel = "output.xlsx"
    
    # 执行操作
    data = read_excel(input_excel)
    generate_xml(data, output_xml)
    # write_excel(data, output_excel)
    append_excel(data, output_excel)
    
    print("处理完成！XML 和 Excel 文件已生成。")
