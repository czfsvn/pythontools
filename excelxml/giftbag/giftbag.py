# -*- coding: GB2312 -*-
"""
Created on Wed Feb 20 18:05:35 2019

@author: Administrator
"""

import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import load_workbook
import copy, os, sys;

class Config:
    def __init__(self):
        self.obj_excel = ""
        self.obj_sheet =  "";
        self.gift_excel =  "";
        self.gift_sheet =  "";
        self.gift_output_xml =  "";
        self.gift_template_xml = "";

class GiftXmlTempate:
    def __init__(self):
        self.root = None;

class NewGift:
    def __init__(self, id):
        self.id = id;
        self.field = [];   # gift���ֶ�
        self.objs = []; 
        self.row = [];
        self.objs = [];
        self.row = [];
        self.xml = None
        self.templateID = 0;
        self.columns = []  # �洢��Ҫ�޸ĵ�������
        self.newdata = [];  # �洢��Ҫ�޸ĵ�������


def indent(elem, level=0):
    i = "\n" + level*"  "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "	"
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
        for elem in elem:
            indent(elem, level+1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i

class ConfigGenerator:
    def __init__(self, config):
        self.config = config; 
        self.gifts=[];
    
    def read_headers_openpyxl(self):
        try:
            wb = load_workbook(self.config.obj_excel)
            ws = wb[self.config.obj_sheet]
            
            # ����Ƿ�Ϊ�ձ�
            if ws.max_row == 0:
                return []
            
            # ���ر�ͷ
            return [cell.value for cell in ws[1]]
        except FileNotFoundError:
            print(f"����[read_headers] �ļ� {self.config.obj_excel}/{self.config.obj_sheet} �����ڣ�")
            input("Press Enter to continue...")
            return []
        except KeyError:
            print(f"����[read_headers] ������ {self.config.obj_excel}/{self.config.obj_sheet} �����ڣ�")
            input("Press Enter to continue...")
            return []
        except Exception as e:
            print(f"����{e}, [read_headers] ������ {self.config.obj_excel}/{self.config.obj_sheet} ������");
            input("Press Enter to continue...")
            return []

    def read_lastrow_openpyxl(self):
        try:
            wb = load_workbook(self.config.obj_excel);
            ws = wb[self.config.obj_sheet];

            last_row_num = ws.max_row;
            last_col_num = ws.max_column;
        
            #print(f"last_row_num: {last_row_num}");
            #print(f"last_col_num: {last_col_num}");
    
            # ��ȡ���һ�е�����
            last_row_data = []
            for cell in ws[last_row_num]:
                last_row_data.append(cell.value)

            #print("lastrow data: ", last_row_data);
            return last_row_data;
        except FileNotFoundError:
            print(f"����[read_lastrow] �ļ� {self.config.obj_excel}/{self.config.obj_sheet} �����ڣ�")
            input("Press Enter to continue...")
            return []
        except KeyError:
            print(f"����[read_lastrow] ������ {self.config.obj_excel}/{self.config.obj_sheet} �����ڣ�")
            input("Press Enter to continue...")
            return []
        except Exception as e:
            print(f"����{e}, [read_lastrow] ������ {self.config.obj_excel}/{self.config.obj_sheet} ������")
            input("Press Enter to continue...")
            return None
        
    
    def read_objexcel_by_id(self, id):
        try:
            wb = load_workbook(self.config.obj_excel)
            ws = wb[self.config.obj_sheet]

            last_row_num = ws.max_row;
            last_col_num = ws.max_column;

            #print(f"last_row_num: {last_row_num}");
            #print(f"last_col_num: {last_col_num}");
        
            for row in ws.iter_rows(min_row=2, values_only=True):
                if int(row[0]) == id:
                    return row;
    
            return [];        
        except FileNotFoundError:
            print(f"����[read_objexcel_by_id] �ļ� {self.config.obj_excel}/{self.config.obj_sheet} �����ڣ�")
            return []
        except KeyError:
            print(f"����[read_objexcel_by_id] ������ {self.config.obj_excel}/{self.config.obj_sheet} �����ڣ�")
            return []
        except Exception as e:
            print(f"����{e}, [read_objexcel_by_id] ������ {self.config.obj_excel}/{self.config.obj_sheet} �����ã�")
            input("Press Enter to continue...")
            return []
        
    def read_gift_template(self):
        try:        
            with open(self.config.gift_template_xml, "r", encoding="GB2312") as file:
                content = file.read()

            root = ET.fromstring(content, parser=ET.XMLParser(encoding="GB2312"))
            #tree = ET.parse(self.config.gift_template_xml)
            #root = tree.getroot()
            config = Config();
            for child in root:
                if child.tag == "bag":
                    self.bagattr = child.attrib;
                    for bagchild in child:
                        if bagchild.tag == "rule":
                            self.rulletype = bagchild.attrib["type"];
        except FileNotFoundError:
            print(f"����[read_gift_template] �ļ� {self.config.gift_template_xml} �����ڣ�")
            input("Press Enter to continue...")
            return
        except KeyError:
            print(f"����[read_gift_template] ������ {self.config.gift_template_xml} �����ڣ�")
            input("Press Enter to continue...")
            return
        except Exception as e:
            print(f"����{e}, [read_gift_template] ������ {self.config.gift_template_xml} ������")
            input("Press Enter to continue...")
            return
        
    def getTemplateData(self, teamplateid, oldid):
        try:
            wb = load_workbook(self.config.obj_excel)
            ws = wb[self.config.obj_sheet]

            last_row_num = ws.max_row;
            last_col_num = ws.max_column;

            teamplatedata = [];
            olddata = [];
            max_id = 0;
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == None:
                    continue;
                id = int(row[0]);
                if id > max_id:
                    max_id = id;
                if id == teamplateid:
                    teamplatedata = list(row);
                if id == oldid:
                    olddata = list(row);                    
    
            return (max_id, teamplatedata, olddata);
        except FileNotFoundError:
            print(f"����[getTemplateData] �ļ� {self.config.obj_excel}/{self.config.obj_sheet} �����ڣ�")
            input("Press Enter to continue...")
            return []
        except KeyError:
            print(f"����[getTemplateData] ������ {self.config.obj_excel}/{self.config.obj_sheet} �����ڣ�")
            input("Press Enter to continue...")
            return []
        except Exception as e:
            print(f"����{e}, [getTemplateData] ������ {self.config.obj_excel}/{self.config.obj_sheet} ������")
            input("Press Enter to continue...")
            return []

    def generate_giftxml_old(self):
        print("generate xml");
        root = ET.Element('Config')
        for gift in self.gifts:
            gift_root = ET.SubElement(root, 'GiftBag')
            gift_root.set("giftid", str(gift.id));
            for row in gift.objs:
                record = ET.SubElement(gift_root, 'Data')
                for col_index, value in enumerate(row):
                    record.set(gift.field[col_index], str(value));
    
        indent(root)
        # ���� XML ��
        tree = ET.ElementTree(root)
        # д�� XML �ļ�
        tree.write(self.config.gift_output_xml, encoding='GB2312', xml_declaration=True)

    def generate_giftxml(self):
        #print("generate xml");
        root = ET.Element('Config')
        for gift in self.gifts:
            bagroot = ET.SubElement(root, 'bag')
            for key, value in self.bagattr.items():
                if key == "id":
                    bagroot.set("id", str(gift.id));
                else:
                    bagroot.set(key, str(value));

            rulenode = ET.SubElement(bagroot, 'rule')
            rulenode.set("type", str(self.rulletype));
            
            for row in gift.objs:
                record = ET.SubElement(bagroot, 'item')
                for col_index, value in enumerate(row):
                    if value == None:
                        continue;
                    record.set(gift.field[col_index], str(value));
            
        indent(root)
        # ���� XML ��
        tree = ET.ElementTree(root)
        # д�� XML �ļ�
        tree.write(self.config.gift_output_xml, encoding='GB2312', xml_declaration=True)


    
    
    def update_old(self, gift):
        try:
            wb = load_workbook(self.config.obj_excel)
            ws = wb[self.config.obj_sheet]

            updated = {};
            log_dict = {};
            for index, headername in enumerate(self.header):
                for col_index, value in enumerate(gift.columns):
                    if headername == value:
                        updated[index + 1] = gift.newdata[col_index];
                        log_dict[headername] = gift.newdata[col_index];

            needsave = False;
            for row in ws.iter_rows(min_row=2):
                findrow = False;
                savedcount = 0;
                for cell in row:
                    if findrow == False and cell.column != 1:
                        continue;
                    
                    if cell.value == gift.id:
                        #print(f"ֵ: {cell.value}")
                        #print(f"����: {cell.coordinate}")  # �� A2��B3
                        #print(f"�к�: {cell.row}, �к�: {cell.column}")
                        #print("---")
                        findrow = True;

                    if findrow == True:
                        if cell.column in updated:
                            cell.value = updated[cell.column];
                            needsave = True;
                            savedcount = savedcount + 1;

                            
                    if savedcount == len(updated):
                        break;
                
                if savedcount == len(updated):
                    break;

            if needsave:
                wb.save(self.config.obj_excel)
                print("[OK]: update old data success, id: ", gift.id)
                print("[OK]: update old data success, updated: ", log_dict);        
    
        except FileNotFoundError:
            print(f"����[update_old] �ļ� {self.config.gift_template_xml} �����ڣ�")
            input("Press Enter to continue...")
            return []
        except KeyError:
            print(f"����[update_old] ������ {self.config.gift_template_xml} �����ڣ�")
            input("Press Enter to continue...")
            return []
        except Exception as e:
            print(f"����{e}, [update_old] ������ {self.config.gift_template_xml} ������")
            input("Press Enter to continue...")
            return []
    
    def fill_gift_row_new(self, gift):
        for index, headername in enumerate(self.header):
            #print(f"���� {index}: {headername}")
            for col_index, value in enumerate(gift.columns):
                if headername == value:
                    gift.row[index] = gift.newdata[col_index];
                    break;

    def fill_gift_row(self, gift):
        data = self.getTemplateData(gift.templateID, gift.id);
        if data == None:
            print("error: getTemplateData failed");
            return;

        max_id = data[0];
        teamplatedata = data[1];
        olddata = data[2];
    
        if gift.id == 0:    # ��Ҫ�����µ���
            if teamplatedata == None or len(teamplatedata) == 0:
                print("error: teamplatedata is None");
                return;
        
            gift.id = max_id + 1;
            gift.row = teamplatedata;
            gift.row[0] = gift.id;
            self.fill_gift_row_new(gift);
            self.write_objexcel(gift);

        elif gift.id > 0:
            if olddata != None and len(olddata) > 0: # ��Ҫ�����ϵ�����Ϣ, ��ʹ��gift.id
                # todo: ���Ǿ�����
                self.update_old(gift);
                return;            
            else:   # ��Ҫ�����µ�����Ϣ�� ��ʹ��gift.id
                if teamplatedata == None or len(teamplatedata) == 0:
                    print("error: teamplatedata is None");
                    return;
                gift.row = teamplatedata;
                gift.row[0] = gift.id;
                self.fill_gift_row_new(gift);
                self.write_objexcel(gift);
    
    
    def write_objexcel(self, gift):
        try:
            if len(gift.row) == 0:
                return;

            wb = load_workbook(self.config.obj_excel)
            ws = wb[self.config.obj_sheet]
            
            #print("write excel, giftid: ", gift.id);
            #print("write excel, giftrow: ", gift.row);

            ws.append(gift.row)
            wb.save(self.config.obj_excel);
        except FileNotFoundError:
            print(f"����[write_objexcel] �ļ� {self.config.obj_excel}/{self.config.obj_sheet}  �����ڣ�")
            input("Press Enter to continue...")
            return []
        except KeyError:
            print(f"����[write_objexcel] ������ {self.config.obj_excel}/{self.config.obj_sheet} �����ڣ�")
            input("Press Enter to continue...")
            return []
        except Exception as e:
            print(f"����{e}, [write_objexcel] ������ {self.config.obj_excel}/{self.config.obj_sheet}  �����ã�")
            input("Press Enter to continue...")
            return []
    
    def process_all_gifts(self):        
        for gift in self.gifts:
            self.fill_gift_row(gift);
        
        print("[%s][%s] [excel]�������" %(self.config.obj_excel, self.config.obj_sheet))
        
        self.generate_giftxml();
        print("[%s][xml]�������" %(self.config.gift_output_xml))
    
    def read_gifts(self):
        try: 
            wb = load_workbook(self.config.gift_excel)
            ws = wb[self.config.gift_sheet]

            last_row_num = ws.max_row;
            last_col_num = ws.max_column;
        
            #print(f"last_row_num: {last_row_num}");
            #print(f"last_col_num: {last_col_num}");

            # ����������
            gift = NewGift(0);
            
            for row in ws.iter_rows(min_row=1, values_only=True):
                if row[0] == None:
                    continue;
                if  row[0] == 'id':
                    if len(gift.field) == 0:
                        gift.field = row;
                elif row[0] == '���' and row[1] == '�µ���':
                    #print("new gift: ��Ҫ�����µĵ���");
                    self.emplace_gift(gift)
                    gift = NewGift(0);
                elif row[0] == '���' and row[1] != None and int(row[1] > 0):
                    #print("new gift: ��Ҫ�����ϵĵ���");
                    self.emplace_gift(gift)
                    gift = NewGift(int(row[1]));
                    continue;
                elif row[0] == 'templateID':
                    gift.columns = row;
                else:
                    if len(gift.columns) != 0 and len(gift.newdata) == 0 and gift.templateID == 0:
                        gift.newdata = row;
                        gift.templateID = int(row[0]);
                    else:
                        gift.objs.append(row);
            
            self.emplace_gift(gift);            
        except FileNotFoundError:
            print(f"����[read_gifts] �ļ� {self.config.gift_excel}/{self.config.gift_sheet} �����ڣ�")
            input("Press Enter to continue...")
            return []
        except KeyError:
            print(f"����[read_gifts] ������ {self.config.gift_excel}/{self.config.gift_sheet} �����ڣ�")
            input("Press Enter to continue...")
            return []
        except Exception as e:
            print(f"����{e}, [read_gifts] ������ {self.config.gift_excel}/{self.config.gift_sheet} ������")
            input("Press Enter to continue...")
            return []

    def run(self):
        #last_row_data = self.read_lastrow();
        #print("last_row_data: ",last_row_data);
        #self.header = self.read_headers();
        #print("headers: ",header);
        #self.read_gift_template();
        self.read_gifts();
        self.process_all_gifts();
        

def read_config(configpath):    
    try:
        with open(configpath, "r", encoding="GB2312") as file:
            content = file.read()

        root = ET.fromstring(content, parser=ET.XMLParser(encoding="GB2312"))
        config = Config();
        for child in root:
            #print(f"tagname={child.tag}, attrib={child.attrib}");
            if child.tag == "objitem":
                config.obj_excel = child.attrib["filepath"];
                config.obj_sheet = child.attrib["sheet"];
            if child.tag == "giftitem":
                config.gift_excel = child.attrib["filepath"]    
                config.gift_sheet = child.attrib["sheet"];
            elif child.tag == "giftxml":
                config.gift_output_xml = child.attrib["outpath"];
                config.gift_template_xml = child.attrib["templatexml"];

        return config;
    except FileNotFoundError:
        print(f"����[read_config] �ļ� {configpath} �����ڣ�")
        input("Press Enter to continue...")
        return None
    except KeyError:
        print(f"����[read_config] ������ {configpath} �����ڣ�")
        input("Press Enter to continue...")
        return None
    except Exception as e:
        print(f"����{e}, [read_config] ������[read_config] {configpath} ������")
        input("Press Enter to continue...")
        return None


if __name__=="__main__":
    base_path = Path(__file__).parent/"config.xml"
    config = read_config(base_path);
    #config = read_config('./config.xml');
    
    generator = ConfigGenerator(config);
    generator.run()
    input("Press Enter to continue...")

